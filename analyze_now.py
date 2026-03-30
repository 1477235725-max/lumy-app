import openpyxl

file_path = 'data.xlsx'
output_file = 'analysis_final.txt'

wb = openpyxl.load_workbook(file_path, data_only=True)

with open(output_file, 'w', encoding='utf-8') as f:
    f.write("="*80 + "\n")
    f.write("地图付费和广告数据分析报告\n")
    f.write("="*80 + "\n\n")
    
    f.write(f"工作表数量: {len(wb.sheetnames)}\n")
    f.write(f"工作表列表: {wb.sheetnames}\n\n")
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        f.write("="*80 + "\n")
        f.write(f"工作表: {sheet_name}\n")
        f.write("="*80 + "\n")
        
        max_row = sheet.max_row
        max_col = sheet.max_column
        f.write(f"数据维度: {max_row} 行 × {max_col} 列\n\n")
        
        f.write("【字段列表】\n")
        headers = []
        for col in range(1, max_col + 1):
            header_value = sheet.cell(row=1, column=col).value
            headers.append(header_value)
            if header_value:
                f.write(f"  {col}. {header_value}\n")
        f.write("\n")
        
        f.write("【数据预览（前10行）】\n")
        for row in range(1, min(11, max_row + 1)):
            row_data = []
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value else '')
            f.write("  " + " | ".join(row_data) + "\n")
        f.write("\n")
        
        f.write("【商业化分析】\n")
        
        payment_keywords = ['付费', '支付', '收入', 'revenue', 'payment', 'pay']
        ad_keywords = ['广告', '投放', '曝光', '点击', 'ctr', 'cpm', 'ad', 'advertisement']
        
        payment_cols = []
        ad_cols = []
        
        for col in range(1, max_col + 1):
            header_value = sheet.cell(row=1, column=col).value
            if header_value:
                header_str = str(header_value).lower()
                if any(keyword in header_str for keyword in payment_keywords):
                    payment_cols.append(header_value)
                if any(keyword in header_str for keyword in ad_keywords):
                    ad_cols.append(header_value)
        
        if payment_cols:
            f.write(f"  发现付费相关字段: {payment_cols}\n")
        else:
            f.write("  未发现明确的付费相关字段\n")
        
        if ad_cols:
            f.write(f"  发现广告相关字段: {ad_cols}\n")
        else:
            f.write("  未发现明确的广告相关字段\n")
        
        f.write("\n")
    
    f.write("="*80 + "\n")
    f.write("分析完成\n")
    f.write("="*80 + "\n")
